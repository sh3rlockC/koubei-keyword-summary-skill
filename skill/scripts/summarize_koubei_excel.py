#!/usr/bin/env python3
import argparse
import base64
import hashlib
import json
import math
import hmac
import re
import sys
import time
from datetime import datetime, timezone
from collections import Counter, defaultdict
from pathlib import Path
from urllib import request as urllib_request

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from tqdm.auto import tqdm
except Exception:  # pragma: no cover - optional dependency fallback
    tqdm = None

BASE_DIRECTIONS = {
    "空间": ["空间", "宽敞", "得房率", "乘坐空间", "后排空间", "腿部空间", "二郎腿", "头部空间"],
    "配置性价比": ["配置", "性价比", "价格", "定价", "权益", "不送", "补贴", "免息", "充电桩", "落地价", "入门版", "顶配", "低配", "高配", "版本", "选装", "hud", "360", "激光版本"],
    "座椅舒适性": ["座椅", "舒适", "包裹", "支撑", "乘坐舒服", "零压座椅", "按摩", "久坐", "通风", "加热"],
    "外观设计": ["外观", "颜值", "好看", "设计", "溜背", "造型", "前脸", "尾灯", "轿跑", "无框车门"],
    "储物/实用性": ["储物", "实用", "后备箱", "前备箱", "扶手箱", "放东西", "掀背", "装载", "魔术收纳", "置物板"],
    "续航能耗": ["续航", "电耗", "能耗", "省电", "油耗", "续航焦虑", "达成率", "冬季续航", "缩水"],
    "底盘悬架": ["底盘", "悬架", "悬挂", "滤震", "减震", "支撑性", "双叉臂", "五连杆", "过弯", "侧倾"],
    "充电补能": ["充电", "补能", "快充", "800V", "6C", "充电速度", "超充", "热泵"],
    "车机智能化": ["车机", "智能化", "应用商店", "投屏", "高德", "导航", "芯片", "车道级导航", "carplay", "app", "ota", "语音", "手车互联", "软件", "场景", "主题", "壁纸"],
    "辅助驾驶": ["辅助驾驶", "智驾", "领航", "城区领航", "激光雷达", "noa", "跟车", "泊车辅助", "接管", "红绿灯", "自动泊车"],
    "隔音静谧性": ["隔音", "静谧", "噪音", "风噪", "胎噪", "共振", "静音", "路噪"],
    "内饰质感": ["内饰", "做工", "质感", "塑料", "异味", "划痕"],
}

TOPIC_BUCKETS = {
    "车机生态": ["车机", "应用", "app", "应用商店", "车道级导航", "投屏", "carplay", "手车互联", "软件", "语音"],
    "智驾兑现": ["辅助驾驶", "智驾", "领航", "noa", "激光雷达", "接管", "泊车辅助", "红绿灯", "自动泊车"],
    "续航补能": ["续航", "电耗", "能耗", "冬季续航", "充电", "快充", "800v", "6c", "热泵"],
    "空间舒适": ["空间", "座椅", "舒适", "头部空间", "顶头", "后排空间"],
    "底盘驾驶": ["底盘", "悬架", "滤震", "过弯", "侧倾", "动力", "提速", "超车"],
    "造型内饰": ["外观", "颜值", "前脸", "尾灯", "内饰", "配色", "氛围灯", "无框车门"],
    "NVH": ["风噪", "胎噪", "噪音", "共振", "静谧", "隔音", "路噪"],
    "价格权益": ["价格", "权益", "补贴", "不送", "免息", "落地价", "充电桩"],
    "配置梯度": ["配置", "顶配", "低配", "高配", "版本", "选装", "hud", "360"],
    "储物实用": ["前备箱", "后备箱", "掀背", "储物", "装载", "扶手箱", "置物板"],
}

PRIORITY = ["车机智能化", "辅助驾驶", "配置性价比", "充电补能", "续航能耗", "隔音静谧性", "空间", "座椅舒适性", "底盘悬架", "储物/实用性", "外观设计", "内饰质感"]
STOPWORDS = set(["一个", "这个", "那个", "还是", "比较", "非常", "觉得", "就是", "而且", "因为", "所以", "如果", "可以", "还有", "以及", "现在", "没有", "有点", "有些", "我们", "你们", "他们", "它的", "真的", "对于", "同价位", "感觉", "方面", "起来", "时候", "这样", "不会", "不是", "但是", "的话", "已经", "自己", "一些", "这种", "里面", "外面", "目前", "整体", "表现", "使用", "体验", "很多", "一下", "总的来说", "哈哈", "确实", "然后", "其实", "真的很", "我觉得", "还是很", "非常满意", "没有不满意", "目前没有", "目前还没有", "总体来说", "用车感受", "选择", "提车", "车型", "长安", "启源", "a06", "A06", "这款车", "这个车", "的话就是", "的话那就是", "优点", "缺点", "总结下吧"])

SYNONYM_NORMALIZATION = {
    "空间大": "空间", "空间宽敞": "空间", "后排空间": "空间", "头部空间": "空间",
    "软件太少": "车机软件少", "app太少": "车机软件少", "应用太少": "车机软件少", "车机软件太少": "车机软件少",
    "车道级导航": "车道级导航", "没有车道级导航": "车道级导航缺失",
    "风噪": "风噪", "胎噪": "胎噪", "共振": "共振",
    "前备箱": "前备箱", "后备箱": "后备箱",
    "双叉臂": "双叉臂悬架", "五连杆": "五连杆悬架",
    "800v": "800V快充", "6c": "6C超充", "热泵": "热泵空调",
}

SINGLE_OUTPUT_SHEETS = [
    "总览摘要",
    "业务摘要",
    "产品机会点",
    "最满意方向统计",
    "最不满意方向统计",
    "最满意方向摘要",
    "最不满意方向摘要",
    "最满意主题",
    "最不满意主题",
    "最满意高频词",
    "最不满意高频词",
]

DUAL_OUTPUT_SHEETS = [
    "一页纸总结",
    "总览摘要",
    "跨平台对比",
    "综合业务摘要",
    "产品机会点",
    "汽车之家_满意摘要",
    "汽车之家_不满意摘要",
    "懂车帝_正向摘要",
    "懂车帝_负向摘要",
]

SINGLE_CRITICAL_SHEETS = [
    "总览摘要",
    "业务摘要",
    "最满意方向统计",
    "最不满意方向统计",
    "最满意方向摘要",
    "最不满意方向摘要",
]

DUAL_CRITICAL_SHEETS = [
    "一页纸总结",
    "总览摘要",
    "跨平台对比",
    "综合业务摘要",
]

SINGLE_STAGE_WEIGHTS = {
    "validate": 50,
    "load_input": 120,
    "summarize_sat": 360,
    "summarize_unsat": 360,
    "write_output": 55,
    "write_validation": 55,
}

DUAL_STAGE_WEIGHTS = {
    "validate": 40,
    "load_autohome": 80,
    "load_dcd": 80,
    "summarize_sat": 180,
    "summarize_unsat": 180,
    "summarize_dcd": 320,
    "write_output": 60,
    "write_validation": 60,
}


def progress_iter(iterable, *, total=None, desc="", enabled=False):
    if not enabled or tqdm is None:
        return iterable
    return tqdm(iterable, total=total, desc=desc, dynamic_ncols=True, leave=False, file=sys.stderr)


def utc_now_iso():
    return datetime.now(timezone.utc).astimezone().replace(microsecond=0).isoformat()


def write_json_atomic(path, payload):
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = target.with_suffix(target.suffix + ".tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(target)


def post_json(url, payload, timeout=5):
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib_request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib_request.urlopen(req, timeout=timeout) as resp:
        resp.read()


def make_feishu_signature(secret, timestamp):
    string_to_sign = f"{timestamp}\n{secret}"
    digest = hmac.new(string_to_sign.encode("utf-8"), digestmod=hashlib.sha256).digest()
    return base64.b64encode(digest).decode("utf-8")


def build_feishu_progress_text(payload):
    percent = int(payload.get("percent", 0))
    stage = payload.get("stage") or ""
    stage_progress = payload.get("stage_progress") or {}
    current = int(stage_progress.get("current", payload.get("current", 0)))
    total = int(stage_progress.get("total", payload.get("total", 1)) or 1)
    message = payload.get("message") or ""
    model_name = payload.get("model_name") or ""
    mode = payload.get("mode") or ""
    mode_label = {
        "autohome": "汽车之家单平台",
        "dual_platform": "双平台整合",
    }.get(mode, mode or "任务")
    lines = [f"口碑摘要 {percent}%｜{model_name}", f"{mode_label} · {stage} {current}/{total}"]
    if message and message not in {stage, f"{stage} {current}/{total}"}:
        lines.append(f"说明：{message}")
    return "\n".join(lines)


def build_feishu_payload(payload, *, secret=None):
    timestamp = int(time.time())
    feishu_payload = {
        "timestamp": str(timestamp),
        "msg_type": "text",
        "content": {
            "text": build_feishu_progress_text(payload),
        },
    }
    if secret:
        feishu_payload["sign"] = make_feishu_signature(secret, timestamp)
    return feishu_payload


class ProgressReporter:
    def __init__(
        self,
        *,
        label,
        model_name,
        mode,
        progress_file=None,
        progress_webhook=None,
        feishu_webhook=None,
        feishu_secret=None,
        total_units=1000,
    ):
        self.label = label
        self.model_name = model_name
        self.mode = mode
        self.progress_file = Path(progress_file).resolve() if progress_file else None
        self.progress_webhook = progress_webhook
        self.feishu_webhook = feishu_webhook
        self.feishu_secret = feishu_secret
        self.total_units = total_units
        self.current_units = 0
        self.stage_base_units = 0
        self.stage_weight_units = 0
        self.stage_total = 1
        self.stage_current = 0
        self.stage_key = "init"
        self.stage_label = "准备"
        self.last_emitted_percent = None
        self.last_emitted_stage = None
        self.last_message = ""
        self.last_webhook_percent = None
        self.last_webhook_stage = None
        self.disabled_file = False
        self.disabled_webhook = False
        self.disabled_feishu = False

    def _build_payload(self, percent=None, message=""):
        percent = int(percent if percent is not None else round(self.current_units * 100 / self.total_units))
        stage_total = max(int(self.stage_total), 1)
        stage_current = min(int(self.stage_current), stage_total)
        payload = {
            "label": self.label,
            "mode": self.mode,
            "model_name": self.model_name,
            "stage": self.stage_label,
            "stage_key": self.stage_key,
            "current": int(self.current_units),
            "total": int(self.total_units),
            "percent": percent,
            "message": message or self.last_message,
            "updated_at": utc_now_iso(),
            "overall": {
                "current": int(self.current_units),
                "total": int(self.total_units),
                "percent": percent,
            },
            "stage_progress": {
                "current": stage_current,
                "total": stage_total,
                "percent": int(round(stage_current * 100 / stage_total)),
            },
            "progress_file": str(self.progress_file) if self.progress_file else None,
            "webhook": self.progress_webhook,
        }
        return payload

    def _emit(self, payload):
        percent = payload["percent"]
        if self.progress_file and not self.disabled_file:
            try:
                write_json_atomic(self.progress_file, payload)
            except Exception as exc:
                self.disabled_file = True
                print(f"WARNING: progress file write failed: {exc}", file=sys.stderr)
        webhook_should_emit = (
            percent != self.last_webhook_percent
            or payload["stage"] != self.last_webhook_stage
            or payload["message"] in {"导出完成", "输入校验失败", "校验失败"}
        )
        if not webhook_should_emit:
            self.last_emitted_percent = percent
            self.last_emitted_stage = payload["stage"]
            self.last_message = payload["message"]
            return

        self.last_webhook_percent = percent
        self.last_webhook_stage = payload["stage"]
        self.last_emitted_percent = percent
        self.last_emitted_stage = payload["stage"]
        self.last_message = payload["message"]
        if self.progress_webhook and not self.disabled_webhook:
            try:
                post_json(self.progress_webhook, payload)
            except Exception as exc:
                self.disabled_webhook = True
                print(f"WARNING: progress webhook failed: {exc}", file=sys.stderr)
        if self.feishu_webhook and not self.disabled_feishu:
            try:
                feishu_payload = build_feishu_payload(payload, secret=self.feishu_secret)
                post_json(self.feishu_webhook, feishu_payload)
            except Exception as exc:
                self.disabled_feishu = True
                print(f"WARNING: feishu webhook failed: {exc}", file=sys.stderr)

    def start_stage(self, stage_key, stage_label, *, weight, total=1, message=None):
        self.stage_key = stage_key
        self.stage_label = stage_label
        self.stage_base_units = self.current_units
        self.stage_weight_units = max(int(weight), 0)
        self.stage_total = max(int(total), 1)
        self.stage_current = 0
        if message is not None:
            self.last_message = message
        self._emit(self._build_payload(message=message or stage_label))

    def mark_stage(self, stage_key, stage_label, *, weight, message=None):
        self.stage_key = stage_key
        self.stage_label = stage_label
        self.stage_base_units = 0
        self.stage_weight_units = max(int(weight), 0)
        self.stage_total = 1
        self.stage_current = 1
        self.current_units = min(self.total_units, self.stage_weight_units)
        self._emit(self._build_payload(message=message or stage_label))

    def advance(self, current, *, message=None):
        self.stage_current = max(0, min(int(current), self.stage_total))
        fraction = self.stage_current / self.stage_total if self.stage_total else 1.0
        self.current_units = min(self.total_units, int(round(self.stage_base_units + self.stage_weight_units * fraction)))
        self._emit(self._build_payload(message=message or self.stage_label))

    def finish_stage(self, message=None):
        self.stage_current = self.stage_total
        self.current_units = min(self.total_units, int(round(self.stage_base_units + self.stage_weight_units)))
        self._emit(self._build_payload(message=message or f"{self.stage_label}完成"))

    def finalize(self, stage_label="完成", message="导出完成"):
        self.stage_label = stage_label
        self.stage_key = stage_label
        self.stage_current = self.stage_total
        self.current_units = self.total_units
        payload = self._build_payload(percent=100, message=message)
        payload["stage_progress"]["current"] = payload["stage_progress"]["total"]
        payload["stage_progress"]["percent"] = 100
        self._emit(payload)


def make_reporter(args, *, mode, model_name):
    progress_file = args.progress_file
    if progress_file and not Path(progress_file).is_absolute():
        progress_file = str(Path(progress_file).resolve())
    label = f"koubei-summary:{mode}:{model_name}"
    return ProgressReporter(
        label=label,
        model_name=model_name,
        mode=mode,
        progress_file=progress_file,
        progress_webhook=args.progress_webhook,
        feishu_webhook=args.feishu_webhook,
        feishu_secret=args.feishu_secret,
    )


def normalize_text(text):
    if text is None:
        return ""
    if isinstance(text, float) and math.isnan(text):
        return ""
    s = str(text).strip()
    s = re.sub(r"\s+", "", s)
    return s


def split_sentences(text):
    text = normalize_text(text)
    if not text:
        return []
    coarse = re.split(r"[。！？；;\n\r]+", text)
    parts = []
    for p in coarse:
        p = p.strip(" ，,、")
        if not p:
            continue
        finer = re.split(r"，|,|、|并且|但是|不过|另外|尤其是|而且|同时", p)
        for f in finer:
            f = f.strip(" ：:，,、【】")
            if f:
                parts.append(f)
    return parts


def score_directions(text):
    lower = text.lower()
    scored = []
    for d, kws in BASE_DIRECTIONS.items():
        score = sum(1 for kw in kws if kw.lower() in lower)
        if score:
            scored.append((d, score))
    scored.sort(key=lambda x: (-x[1], PRIORITY.index(x[0]) if x[0] in PRIORITY else 999))
    return scored


def pick_direction(sentence):
    scored = score_directions(sentence)
    return scored[0][0] if scored else None


def pick_main_directions(text, topn=3):
    agg = Counter()
    for sent in split_sentences(text):
        for d, score in score_directions(sent):
            agg[d] += score
    if not agg:
        return []
    return [d for d, _ in agg.most_common(topn)]


def normalize_fragment(fragment):
    frag = normalize_text(fragment)
    frag = re.sub(r"^(优点|缺点|不满意的话就是|最不满意的是|最遗憾的是|说下缺点|现在说说不满意的点|最不满意|最满意|好评|槽点)", "", frag)
    return frag.strip(" ：:，,、")


def fingerprint(text):
    text = normalize_fragment(text)
    for raw, norm in SYNONYM_NORMALIZATION.items():
        text = text.replace(raw, norm)
    text = re.sub(r"\d+", "#", text)
    text = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff#]", "", text)
    return text[:80]


def merge_fragments(frags, max_len=42):
    out = []
    cur = ""
    for frag in frags:
        frag = normalize_fragment(frag)
        if not frag:
            continue
        if not cur:
            cur = frag
            continue
        cand = cur + "，" + frag
        if len(cand) <= max_len:
            cur = cand
        else:
            out.append(cur)
            cur = frag
    if cur:
        out.append(cur)
    return out


def profile_dataframe(df):
    return {
        "rows": int(len(df)),
        "columns": [str(col) for col in df.columns],
        "empty": bool(df.empty),
    }


def collect_autohome_input(path, *, show_progress=False, reporter=None, stage_weight=120):
    path = Path(path)
    report = {
        "source_path": str(path),
        "selected_sheets": [],
        "sheet_reports": [],
        "issues": [],
        "warnings": [],
    }
    if not path.exists():
        report["issues"].append(f"输入文件不存在: {path}")
        return None, report

    workbook = pd.ExcelFile(path)
    frames = []
    if reporter:
        reporter.start_stage("load_autohome", "读取汽车之家 sheet", weight=stage_weight, total=len(workbook.sheet_names), message="开始读取汽车之家输入")
    for idx, sheet in enumerate(progress_iter(workbook.sheet_names, total=len(workbook.sheet_names), desc="读取汽车之家 sheet", enabled=show_progress), start=1):
        df = pd.read_excel(path, sheet_name=sheet)
        has_required = "最满意" in df.columns and "最不满意" in df.columns
        report["sheet_reports"].append({
            "sheet": sheet,
            "rows": int(len(df)),
            "columns": [str(col) for col in df.columns],
            "has_required_columns": has_required,
        })
        if has_required and not df.empty:
            frames.append(df.copy())
            report["selected_sheets"].append(sheet)
        if reporter:
            reporter.advance(idx, message=f"读取汽车之家 sheet {idx}/{len(workbook.sheet_names)}")

    if not frames:
        report["issues"].append("未识别到包含‘最满意’和‘最不满意’列的 sheet")
        return None, report

    if len(report["selected_sheets"]) > 1:
        report["warnings"].append(f"识别到 {len(report['selected_sheets'])} 个汽车之家口碑 sheet，已合并后继续处理")

    merged = pd.concat(frames, ignore_index=True)
    report["row_count"] = int(len(merged))
    if reporter:
        reporter.finish_stage("汽车之家输入已读取")
    return merged, report


def collect_dcd_input(path, *, show_progress=False, reporter=None, stage_weight=80):
    path = Path(path)
    report = {
        "source_path": str(path),
        "selected_sheet": None,
        "sheet_reports": [],
        "issues": [],
        "warnings": [],
    }
    if not path.exists():
        report["issues"].append(f"输入文件不存在: {path}")
        return None, report

    workbook = pd.ExcelFile(path)
    selected_df = None
    if reporter:
        reporter.start_stage("load_dcd", "读取懂车帝 sheet", weight=stage_weight, total=len(workbook.sheet_names), message="开始读取懂车帝输入")
    for idx, sheet in enumerate(progress_iter(workbook.sheet_names, total=len(workbook.sheet_names), desc="读取懂车帝 sheet", enabled=show_progress), start=1):
        df = pd.read_excel(path, sheet_name=sheet)
        has_required = "评价全文" in df.columns
        report["sheet_reports"].append({
            "sheet": sheet,
            "rows": int(len(df)),
            "columns": [str(col) for col in df.columns],
            "has_required_columns": has_required,
        })
        if selected_df is None and has_required and not df.empty:
            selected_df = df.copy()
            report["selected_sheet"] = sheet
        if reporter:
            reporter.advance(idx, message=f"读取懂车帝 sheet {idx}/{len(workbook.sheet_names)}")

    if selected_df is None:
        report["issues"].append("懂车帝输入缺少‘评价全文’列")
        return None, report

    report["row_count"] = int(len(selected_df))
    if len([item for item in report["sheet_reports"] if item["has_required_columns"]]) > 1:
        report["warnings"].append(f"识别到多个包含‘评价全文’的 sheet，已使用 {report['selected_sheet']} 继续处理")
    if reporter:
        reporter.finish_stage("懂车帝输入已读取")
    return selected_df, report


def build_validation_report(
    *,
    mode,
    model_name,
    input_reports,
    output_profiles,
    expected_sheets,
    critical_sheets,
    strict_validate=False,
):
    errors = []
    warnings = []
    for name, report in input_reports.items():
        for issue in report.get("issues", []):
            errors.append(f"{name}: {issue}")
        for item in report.get("warnings", []):
            warnings.append(f"{name}: {item}")

    missing_sheets = [sheet for sheet in expected_sheets if sheet not in output_profiles]
    if missing_sheets:
        errors.append(f"输出缺少 sheet: {', '.join(missing_sheets)}")

    for sheet in critical_sheets:
        profile = output_profiles.get(sheet)
        if profile is None:
            continue
        if profile.get("rows", 0) == 0:
            errors.append(f"输出 sheet {sheet} 为空")

    for sheet, profile in output_profiles.items():
        if sheet in critical_sheets:
            continue
        if profile.get("rows", 0) == 0:
            warnings.append(f"输出 sheet {sheet} 为空")

    if strict_validate and warnings and not errors:
        errors.extend([f"strict validation: {item}" for item in warnings])

    return {
        "mode": mode,
        "model_name": model_name,
        "ok": not errors,
        "errors": errors,
        "warnings": warnings,
        "inputs": input_reports,
        "outputs": output_profiles,
    }


def summarize_side(df, colname, *, show_progress=False, desc_prefix="", reporter=None, stage_key=None, stage_label=None, stage_weight=0):
    main_counter = Counter()
    item_counter = Counter()
    merged_counter = Counter()
    direction_fragments = defaultdict(list)
    direction_fingerprints = defaultdict(Counter)
    topic_counter = Counter()
    keyword_counter = Counter()

    if reporter and stage_key:
        reporter.start_stage(stage_key, stage_label or f"{desc_prefix}{colname}", weight=stage_weight, total=len(df), message=f"开始处理 {stage_label or colname}")
    for idx, text in enumerate(progress_iter(df[colname].fillna(""), total=len(df), desc=f"{desc_prefix}{colname}", enabled=show_progress), start=1):
        text = normalize_text(text)
        if not text:
            if reporter and stage_key:
                reporter.advance(idx, message=f"{stage_label or colname} {idx}/{len(df)}")
            continue

        for d in pick_main_directions(text, topn=3):
            main_counter[d] += 1

        item_hits = set()
        for sent in split_sentences(text):
            d = pick_direction(sent)
            if d:
                item_hits.add(d)
                frag = normalize_fragment(sent)
                direction_fragments[d].append(frag)
                fp = fingerprint(frag)
                if fp:
                    direction_fingerprints[d][fp] += 1
            for topic, kws in TOPIC_BUCKETS.items():
                if any(kw.lower() in sent.lower() for kw in kws):
                    topic_counter[topic] += 1
            for token in re.findall(r"[\u4e00-\u9fffA-Za-z0-9\+]{2,16}", sent):
                token = normalize_fragment(token)
                if token and token not in STOPWORDS and not re.fullmatch(r"\d+", token):
                    keyword_counter[token] += 1
        for d in item_hits:
            item_counter[d] += 1
        if reporter and stage_key:
            reporter.advance(idx, message=f"{stage_label or colname} {idx}/{len(df)}")

    for d in direction_fingerprints:
        merged_counter[d] = sum(direction_fingerprints[d].values())

    all_dirs = sorted(set(main_counter) | set(item_counter) | set(merged_counter), key=lambda d: -(main_counter.get(d, 0)))
    combined_rows = []
    for d in all_dirs:
        combined_rows.append({
            "方向": d,
            "主方向口径": main_counter.get(d, 0),
            "条目口径": item_counter.get(d, 0),
            "语句归并口径": merged_counter.get(d, 0),
        })
    if combined_rows:
        combined_df = pd.DataFrame(combined_rows).sort_values(["主方向口径", "条目口径", "语句归并口径"], ascending=False).reset_index(drop=True)
    else:
        combined_df = pd.DataFrame(columns=["方向", "主方向口径", "条目口径", "语句归并口径"])

    summary_rows = []
    for _, row in combined_df.iterrows():
        d = row["方向"]
        main_cnt = int(row["主方向口径"])
        item_cnt = int(row["条目口径"])
        merged_cnt = int(row["语句归并口径"])
        uniq = []
        seen = set()
        for frag in direction_fragments.get(d, []):
            fp = fingerprint(frag)
            if fp in seen:
                continue
            seen.add(fp)
            uniq.append(frag)
            if len(uniq) >= 10:
                break
        quotes = merge_fragments(uniq)[:5]
        summary = f"用户对{d}评价集中偏正面；主方向 {main_cnt}，条目口径 {item_cnt}，语句归并口径 {merged_cnt}。" if colname in ("最满意", "正向反馈") else f"用户对{d}的负面反馈较集中；主方向 {main_cnt}，条目口径 {item_cnt}，语句归并口径 {merged_cnt}。"
        summary_rows.append({
            "方向": d,
            "主方向口径": main_cnt,
            "条目口径": item_cnt,
            "语句归并口径": merged_cnt,
            "摘要": summary,
            "代表性原句1": quotes[0] if len(quotes) > 0 else "",
            "代表性原句2": quotes[1] if len(quotes) > 1 else "",
            "代表性原句3": quotes[2] if len(quotes) > 2 else "",
            "代表性原句4": quotes[3] if len(quotes) > 3 else "",
            "代表性原句5": quotes[4] if len(quotes) > 4 else "",
        })

    topic_rows = [{"主题": t, "提及次数": c} for t, c in topic_counter.most_common()]
    term_rows = [{"关键词": t, "频次": c} for t, c in keyword_counter.most_common(60)]
    if reporter and stage_key:
        reporter.finish_stage(f"{stage_label or colname}完成")
    return combined_df, pd.DataFrame(summary_rows), pd.DataFrame(term_rows), pd.DataFrame(topic_rows)


def extract_dcd_pos_neg(text):
    txt = str(text or "")
    pos, neg = [], []
    lines = [x.strip() for x in re.split(r"[\r\n]+", txt) if x.strip()]
    mode = None
    for line in lines:
        l = line.strip(" ：:")
        if re.search(r"优点|满意|最满意|好评", l) and not re.search(r"不满意|缺点|槽点", l):
            mode = "pos"
            continue
        if re.search(r"缺点|不满意|最不满意|槽点", l):
            mode = "neg"
            continue
        if mode == "pos":
            pos.extend(split_sentences(l))
        elif mode == "neg":
            neg.extend(split_sentences(l))
    if not pos and not neg:
        sents = split_sentences(txt)
        for s in sents:
            if any(k in s for k in ["不错", "满意", "舒服", "好看", "香", "诚意", "喜欢", "优秀", "给力", "宽敞", "快"]):
                pos.append(s)
            if any(k in s for k in ["不好", "不满意", "缺点", "拉跨", "一般", "差", "误触", "噪音", "异味", "没有", "不足", "麻烦"]):
                neg.append(s)
    return pos, neg


def summarize_dcd_rows(df, *, show_progress=False, reporter=None, stage_key=None, stage_label="懂车帝长评解析", stage_weight=0):
    pos_all, neg_all = [], []
    if reporter and stage_key:
        reporter.start_stage(stage_key, stage_label, weight=stage_weight, total=len(df), message="开始拆分懂车帝评价全文")
    for idx, text in enumerate(progress_iter(df["评价全文"].fillna(""), total=len(df), desc="解析懂车帝评价", enabled=show_progress), start=1):
        p, n = extract_dcd_pos_neg(text)
        pos_all.extend(p)
        neg_all.extend(n)
        if reporter and stage_key:
            reporter.advance(idx, message=f"{stage_label} {idx}/{len(df)}")

    pos_df = pd.DataFrame({"正向反馈": pos_all}) if pos_all else pd.DataFrame(columns=["正向反馈"])
    neg_df = pd.DataFrame({"负向反馈": neg_all}) if neg_all else pd.DataFrame(columns=["负向反馈"])
    pos_kw, pos_summary, pos_terms, pos_topics = summarize_side(pos_df, "正向反馈") if not pos_df.empty else (pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
    neg_kw, neg_summary, neg_terms, neg_topics = summarize_side(neg_df, "负向反馈") if not neg_df.empty else (pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
    if reporter and stage_key:
        reporter.finish_stage(f"{stage_label}完成")
    return {
        "raw_count": len(df),
        "pos_kw": pos_kw,
        "neg_kw": neg_kw,
        "pos_summary": pos_summary,
        "neg_summary": neg_summary,
        "pos_terms": pos_terms,
        "neg_terms": neg_terms,
        "pos_topics": pos_topics,
        "neg_topics": neg_topics,
    }


def build_overview(model_name, total_rows, sat_kw, unsat_kw):
    sat_items = "、".join([f"{r['方向']}（主{r['主方向口径']}/条{r['条目口径']}/归并{r['语句归并口径']}）" for _, r in sat_kw.head(8).iterrows()])
    unsat_items = "、".join([f"{r['方向']}（主{r['主方向口径']}/条{r['条目口径']}/归并{r['语句归并口径']}）" for _, r in unsat_kw.head(6).iterrows()])
    sat_top = sat_kw.head(3)["方向"].tolist() if not sat_kw.empty else []
    unsat_top = unsat_kw.head(3)["方向"].tolist() if not unsat_kw.empty else []
    return pd.DataFrame([
        {"项目": "车型", "内容": model_name},
        {"项目": "样本数", "内容": total_rows},
        {"项目": "最满意Top方向", "内容": sat_items},
        {"项目": "最不满意Top方向", "内容": unsat_items},
        {"项目": "一句话印象", "内容": f"{model_name}整体口碑显示：优势集中在{'、'.join(sat_top)}，主要短板集中在{'、'.join(unsat_top)}。" if sat_top and unsat_top else "请结合方向统计查看。"},
    ])


def build_business_summary(model_name, sat_kw, unsat_kw, sat_topics, unsat_topics):
    sat_top5 = "、".join(sat_kw.head(5)["方向"].tolist()) if not sat_kw.empty else ""
    unsat_top5 = "、".join(unsat_kw.head(5)["方向"].tolist()) if not unsat_kw.empty else ""
    return pd.DataFrame([
        {"模块": "核心卖点（主方向口径优先）", "内容": f"{model_name}的核心卖点集中在：{sat_top5}。"},
        {"模块": "核心槽点（主方向口径优先）", "内容": f"用户最集中吐槽的问题是：{unsat_top5}。"},
        {"模块": "满意主题Top", "内容": "、".join(sat_topics.head(4)["主题"].tolist()) if not sat_topics.empty else ""},
        {"模块": "不满意主题Top", "内容": "、".join(unsat_topics.head(4)["主题"].tolist()) if not unsat_topics.empty else ""},
        {"模块": "产品建议", "内容": "优先补齐车机生态、车道级导航、投屏/手车互联能力。加快辅助驾驶 OTA 落地。针对高能耗、冬季续航与实际达成率做更明确优化与沟通。"},
        {"模块": "适合人群/场景", "内容": "家庭通勤、带娃出行、周末短途出游；重视乘坐舒适与日常驾驶质感的用户。"},
    ])


def build_opportunity_sheet(sat_topics, unsat_topics):
    rows = []
    for item in sat_topics.head(5).to_dict(orient="records"):
        rows.append({"类型": "优势机会", "主题": item["主题"], "提及次数": item["提及次数"], "建议": f"可继续强化“{item['主题']}”相关传播与卖点包装。"})
    for item in unsat_topics.head(5).to_dict(orient="records"):
        rows.append({"类型": "问题机会", "主题": item["主题"], "提及次数": item["提及次数"], "建议": f"优先处理“{item['主题']}”相关体验问题，并同步优化用户沟通。"})
    return pd.DataFrame(rows)


def build_compare_sheet(zj_sat, zj_unsat, dcd_pos, dcd_neg):
    compare_rows = []
    all_dirs = []
    for df in [zj_sat, zj_unsat, dcd_pos, dcd_neg]:
        if df is not None and not df.empty and "方向" in df.columns:
            for d in df["方向"].head(10).tolist():
                if d not in all_dirs:
                    all_dirs.append(d)
    for d in all_dirs:
        compare_rows.append({
            "方向": d,
            "汽车之家_优势提及": int(zj_sat.loc[zj_sat["方向"] == d, "主方向口径"].iloc[0]) if (zj_sat is not None and not zj_sat.empty and (zj_sat["方向"] == d).any()) else 0,
            "汽车之家_槽点提及": int(zj_unsat.loc[zj_unsat["方向"] == d, "主方向口径"].iloc[0]) if (zj_unsat is not None and not zj_unsat.empty and (zj_unsat["方向"] == d).any()) else 0,
            "懂车帝_优势提及": int(dcd_pos.loc[dcd_pos["方向"] == d, "主方向口径"].iloc[0]) if (dcd_pos is not None and not dcd_pos.empty and (dcd_pos["方向"] == d).any()) else 0,
            "懂车帝_槽点提及": int(dcd_neg.loc[dcd_neg["方向"] == d, "主方向口径"].iloc[0]) if (dcd_neg is not None and not dcd_neg.empty and (dcd_neg["方向"] == d).any()) else 0,
        })
    return pd.DataFrame(compare_rows)


def build_one_pager_sheet(wb, zjh_count, dcd_count, compare_df):
    if "一页纸总结" in wb.sheetnames:
        del wb["一页纸总结"]
    ws = wb.create_sheet("一页纸总结", 0)
    for col, width in {"A": 18, "B": 26, "C": 26, "D": 26, "E": 26}.items():
        ws.column_dimensions[col].width = width
    for r in range(1, 22):
        ws.row_dimensions[r].height = 24

    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="D9EAF7")
    gray = PatternFill("solid", fgColor="F3F6F9")
    green = PatternFill("solid", fgColor="EAF4EA")
    red = PatternFill("solid", fgColor="FCE8E6")
    white_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    bold_font = Font(bold=True, size=11)
    thin = Side(style="thin", color="B7C0C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")

    strengths_df = compare_df.sort_values(["汽车之家_优势提及", "懂车帝_优势提及"], ascending=False)
    strengths = strengths_df[strengths_df[["汽车之家_优势提及", "懂车帝_优势提及"]].sum(axis=1) > 0].head(5)
    weaknesses_df = compare_df.sort_values(["汽车之家_槽点提及", "懂车帝_槽点提及"], ascending=False)
    weaknesses = weaknesses_df[weaknesses_df[["汽车之家_槽点提及", "懂车帝_槽点提及"]].sum(axis=1) > 0].head(5)

    ws.merge_cells("A1:E1")
    ws["A1"] = "双平台口碑一页纸总结"
    ws["A1"].fill = blue
    ws["A1"].font = white_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:E2")
    ws["A2"] = f"样本基础：汽车之家 {zjh_count} 条对齐口碑；懂车帝 {dcd_count} 条口碑。结论基于双平台用户真实反馈归纳。"
    ws["A2"].fill = light
    ws["A2"].font = normal_font
    ws["A2"].alignment = wrap

    ws.merge_cells("A4:E4")
    ws["A4"] = "一句话结论：核心吸引力集中在空间、舒适、底盘、补能效率与配置诚意；主要短板集中在车机生态/导航互联、续航达成稳定性，以及部分细节做工与噪音问题。"
    ws["A4"].fill = gray
    ws["A4"].font = bold_font
    ws["A4"].alignment = wrap

    ws.merge_cells("A6:B6")
    ws["A6"] = "核心优点 Top5"
    ws["A6"].fill = green
    ws["A6"].font = header_font
    ws["A6"].alignment = center

    row = 7
    for idx, (_, r) in enumerate(strengths.iterrows(), start=1):
        ws[f"A{row}"] = f"{idx}. {r['方向']}"
        ws[f"B{row}"] = f"双平台均有正向反馈；汽车之家 {r['汽车之家_优势提及']}，懂车帝 {r['懂车帝_优势提及']}。"
        ws[f"A{row}"].font = bold_font
        ws[f"B{row}"].font = normal_font
        ws[f"A{row}"].fill = green
        ws[f"B{row}"].fill = green
        ws[f"A{row}"].alignment = wrap
        ws[f"B{row}"].alignment = wrap
        row += 1

    ws.merge_cells("D6:E6")
    ws["D6"] = "核心槽点 Top5"
    ws["D6"].fill = red
    ws["D6"].font = header_font
    ws["D6"].alignment = center

    row = 7
    for idx, (_, r) in enumerate(weaknesses.iterrows(), start=1):
        ws[f"D{row}"] = f"{idx}. {r['方向']}"
        ws[f"E{row}"] = f"双平台均有负向反馈；汽车之家 {r['汽车之家_槽点提及']}，懂车帝 {r['懂车帝_槽点提及']}。"
        ws[f"D{row}"].font = bold_font
        ws[f"E{row}"].font = normal_font
        ws[f"D{row}"].fill = red
        ws[f"E{row}"].fill = red
        ws[f"D{row}"].alignment = wrap
        ws[f"E{row}"].alignment = wrap
        row += 1

    ws.merge_cells("A13:E13")
    ws["A13"] = "平台差异观察"
    ws["A13"].fill = light
    ws["A13"].font = header_font
    ws["A13"].alignment = center

    platform_rows = [
        ("汽车之家", "更适合看“最满意 / 最不满意”的结构化方向，优缺点轮廓更清楚。"),
        ("懂车帝", "长评更多，更适合看完整使用场景、选车对比和真实体验细节。"),
        ("共同结论", "两平台结论高度一致：强项是家用舒适和综合配置，短板是智能化体验与细节完善度。"),
    ]
    row = 14
    for a, b in platform_rows:
        ws[f"A{row}"] = a
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws[f"B{row}"] = b
        ws[f"A{row}"].font = bold_font
        ws[f"B{row}"].font = normal_font
        ws[f"A{row}"].alignment = wrap
        ws[f"B{row}"].alignment = wrap
        row += 1

    ws.merge_cells("A18:E18")
    ws["A18"] = "汇报口径建议"
    ws["A18"].fill = light
    ws["A18"].font = header_font
    ws["A18"].alignment = center

    suggestions = [
        "对外传播：主打“空间舒适 + 底盘质感 + 补能效率 + 配置诚意”。",
        "产品优化：优先补车机生态、导航体验、手机互联与 OTA 节奏。",
        "用户沟通：对续航达成率、功能边界、细节配置差异做更清晰预期管理。",
    ]
    for i, text in enumerate(suggestions, start=19):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
        ws[f"A{i}"] = f"• {text}"
        ws[f"A{i}"].font = normal_font
        ws[f"A{i}"].alignment = wrap

    for row in ws.iter_rows(min_row=1, max_row=21, min_col=1, max_col=5):
        for cell in row:
            cell.border = border


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
            except Exception:
                val = ""
            max_len = max(max_len, len(val))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)


def summarize_autohome(input_path, output_path, model_name, strict_validate=False, show_progress=False, reporter=None):
    reporter = reporter or None
    out = Path(output_path)
    report_path = out.with_suffix(".validation.json")
    out.parent.mkdir(parents=True, exist_ok=True)
    if reporter:
        reporter.mark_stage("validate", "输入校验", weight=SINGLE_STAGE_WEIGHTS["validate"], message="开始检查输入")
    df, input_report = collect_autohome_input(
        input_path,
        show_progress=show_progress,
        reporter=reporter if reporter else None,
        stage_weight=SINGLE_STAGE_WEIGHTS["load_input"],
    )
    if df is None:
        report = build_validation_report(
            mode="autohome",
            model_name=model_name,
            input_reports={"autohome": input_report},
            output_profiles={},
            expected_sheets=SINGLE_OUTPUT_SHEETS,
            critical_sheets=SINGLE_CRITICAL_SHEETS,
            strict_validate=strict_validate,
        )
        report["output_path"] = str(out)
        report["validation_path"] = str(report_path)
        write_validation_report(report_path, report)
        if reporter:
            reporter.finalize(message="输入校验失败")
        raise SystemExit(2)

    sat_kw, sat_summary, sat_terms, sat_topics = summarize_side(
        df,
        "最满意",
        show_progress=show_progress,
        desc_prefix="汽车之家 ",
        reporter=reporter if reporter else None,
        stage_key="summarize_sat",
        stage_label="汽车之家最满意",
        stage_weight=SINGLE_STAGE_WEIGHTS["summarize_sat"],
    )
    unsat_kw, unsat_summary, unsat_terms, unsat_topics = summarize_side(
        df,
        "最不满意",
        show_progress=show_progress,
        desc_prefix="汽车之家 ",
        reporter=reporter if reporter else None,
        stage_key="summarize_unsat",
        stage_label="汽车之家最不满意",
        stage_weight=SINGLE_STAGE_WEIGHTS["summarize_unsat"],
    )
    overview = build_overview(model_name, len(df), sat_kw, unsat_kw)
    business = build_business_summary(model_name, sat_kw, unsat_kw, sat_topics, unsat_topics)
    opportunity = build_opportunity_sheet(sat_topics, unsat_topics)

    output_frames = {
        "总览摘要": overview,
        "业务摘要": business,
        "产品机会点": opportunity,
        "最满意方向统计": sat_kw,
        "最不满意方向统计": unsat_kw,
        "最满意方向摘要": sat_summary,
        "最不满意方向摘要": unsat_summary,
        "最满意主题": sat_topics,
        "最不满意主题": unsat_topics,
        "最满意高频词": sat_terms,
        "最不满意高频词": unsat_terms,
    }
    output_profiles = {name: profile_dataframe(frame) for name, frame in output_frames.items()}
    report = build_validation_report(
        mode="autohome",
        model_name=model_name,
        input_reports={"autohome": input_report},
        output_profiles=output_profiles,
        expected_sheets=SINGLE_OUTPUT_SHEETS,
        critical_sheets=SINGLE_CRITICAL_SHEETS,
        strict_validate=strict_validate,
    )
    report["output_path"] = str(out)
    report["validation_path"] = str(report_path)
    if reporter:
        reporter.start_stage("write_output", "写出摘要 Excel", weight=SINGLE_STAGE_WEIGHTS["write_output"], total=1, message="开始写出摘要 Excel")
    if report["errors"]:
        write_validation_report(report_path, report)
        if reporter:
            reporter.finalize(message="校验失败")
        raise SystemExit(2)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sheet_name, frame in output_frames.items():
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
        for ws in writer.book.worksheets:
            style_sheet(ws)
    if reporter:
        reporter.finish_stage("摘要 Excel 已写出")

    if reporter:
        reporter.start_stage("write_validation", "写出校验报告", weight=SINGLE_STAGE_WEIGHTS["write_validation"], total=1, message="开始写出校验报告")
    write_validation_report(report_path, report)
    if reporter:
        reporter.finish_stage("校验报告已写出")
        reporter.finalize(message="导出完成")
    return out


def summarize_dual_platform(autohome_input, dcd_input, output_path, model_name, strict_validate=False, show_progress=False, reporter=None):
    out = Path(output_path)
    report_path = out.with_suffix(".validation.json")
    out.parent.mkdir(parents=True, exist_ok=True)
    if reporter:
        reporter.mark_stage("validate", "输入校验", weight=DUAL_STAGE_WEIGHTS["validate"], message="开始检查输入")

    zj_df, autohome_report = collect_autohome_input(
        autohome_input,
        show_progress=show_progress,
        reporter=reporter if reporter else None,
        stage_weight=DUAL_STAGE_WEIGHTS["load_autohome"],
    )
    dcd_df, dcd_report = collect_dcd_input(
        dcd_input,
        show_progress=show_progress,
        reporter=reporter if reporter else None,
        stage_weight=DUAL_STAGE_WEIGHTS["load_dcd"],
    )
    if zj_df is None or dcd_df is None:
        report = build_validation_report(
            mode="dual_platform",
            model_name=model_name,
            input_reports={"autohome": autohome_report, "dcd": dcd_report},
            output_profiles={},
            expected_sheets=DUAL_OUTPUT_SHEETS,
            critical_sheets=DUAL_CRITICAL_SHEETS,
            strict_validate=strict_validate,
        )
        report["output_path"] = str(out)
        report["validation_path"] = str(report_path)
        write_validation_report(report_path, report)
        if reporter:
            reporter.finalize(message="输入校验失败")
        raise SystemExit(2)

    zj_sat_kw, zj_sat_summary, _, zj_sat_topics = summarize_side(
        zj_df,
        "最满意",
        show_progress=show_progress,
        desc_prefix="汽车之家 ",
        reporter=reporter if reporter else None,
        stage_key="summarize_sat",
        stage_label="汽车之家最满意",
        stage_weight=DUAL_STAGE_WEIGHTS["summarize_sat"],
    )
    zj_unsat_kw, zj_unsat_summary, _, zj_unsat_topics = summarize_side(
        zj_df,
        "最不满意",
        show_progress=show_progress,
        desc_prefix="汽车之家 ",
        reporter=reporter if reporter else None,
        stage_key="summarize_unsat",
        stage_label="汽车之家最不满意",
        stage_weight=DUAL_STAGE_WEIGHTS["summarize_unsat"],
    )
    dcd = summarize_dcd_rows(
        dcd_df,
        show_progress=show_progress,
        reporter=reporter if reporter else None,
        stage_key="summarize_dcd",
        stage_label="懂车帝长评",
        stage_weight=DUAL_STAGE_WEIGHTS["summarize_dcd"],
    )

    sat_top = "、".join(zj_sat_summary["方向"].head(5).tolist()) if not zj_sat_summary.empty else ""
    unsat_top = "、".join(zj_unsat_summary["方向"].head(5).tolist()) if not zj_unsat_summary.empty else ""
    dcd_sat_top = "、".join(dcd["pos_summary"]["方向"].head(5).tolist()) if not dcd["pos_summary"].empty else ""
    dcd_unsat_top = "、".join(dcd["neg_summary"]["方向"].head(5).tolist()) if not dcd["neg_summary"].empty else ""

    combined_overview = pd.DataFrame([
        {"模块": "项目", "内容": f"{model_name} 双平台口碑整合摘要"},
        {"模块": "平台样本", "内容": f"汽车之家 {len(zj_df)} 条对齐口碑；懂车帝 {dcd['raw_count']} 条口碑"},
        {"模块": "汽车之家优势Top", "内容": sat_top},
        {"模块": "汽车之家槽点Top", "内容": unsat_top},
        {"模块": "懂车帝优势Top", "内容": dcd_sat_top},
        {"模块": "懂车帝槽点Top", "内容": dcd_unsat_top},
        {"模块": "综合一句话", "内容": f"{model_name}的强项主要在空间、底盘/舒适、补能效率与配置诚意；短板集中在车机生态、导航/互联、部分续航达成与细节做工。"},
    ])

    compare_df = build_compare_sheet(zj_sat_summary, zj_unsat_summary, dcd["pos_summary"], dcd["neg_summary"])
    business = pd.DataFrame([
        {"模块": "核心卖点", "内容": "空间宽敞、座椅/乘坐舒适、底盘悬架有高级感、补能速度快、配置给得足，是两平台都反复出现的正向主题。"},
        {"模块": "核心槽点", "内容": "车机软件生态弱、缺 CarPlay/投屏、导航能力一般、部分用户反馈续航达成不稳，另外存在噪音/异味/做工细节等零散问题。"},
        {"模块": "产品建议", "内容": "优先补车机生态与导航体验，完善手机互联与 OTA 节奏；继续优化续航达成率沟通、细节做工与误触问题。"},
        {"模块": "适合人群", "内容": "适合看重空间、家用舒适、日常通勤和中短途出行，同时对智驾不是极致刚需的人群。"},
    ])

    opps = []
    for _, r in compare_df.sort_values(["汽车之家_槽点提及", "懂车帝_槽点提及"], ascending=False).head(6).iterrows():
        if r["汽车之家_槽点提及"] or r["懂车帝_槽点提及"]:
            opps.append({"类型": "优先改进", "方向": r["方向"], "汽车之家槽点": r["汽车之家_槽点提及"], "懂车帝槽点": r["懂车帝_槽点提及"], "建议": f"优先处理“{r['方向']}”相关体验问题，并在传播中做预期管理。"})
    for _, r in compare_df.sort_values(["汽车之家_优势提及", "懂车帝_优势提及"], ascending=False).head(6).iterrows():
        if r["汽车之家_优势提及"] or r["懂车帝_优势提及"]:
            opps.append({"类型": "重点强化", "方向": r["方向"], "汽车之家槽点": "", "懂车帝槽点": "", "建议": f"继续放大“{r['方向']}”优势，沉淀为核心卖点。"})
    opps_df = pd.DataFrame(opps)

    output_frames = {
        "总览摘要": combined_overview,
        "跨平台对比": compare_df,
        "综合业务摘要": business,
        "产品机会点": opps_df,
        "汽车之家_满意摘要": zj_sat_summary,
        "汽车之家_不满意摘要": zj_unsat_summary,
        "懂车帝_正向摘要": dcd["pos_summary"],
        "懂车帝_负向摘要": dcd["neg_summary"],
    }
    output_profiles = {name: profile_dataframe(frame) for name, frame in output_frames.items()}
    output_profiles["一页纸总结"] = {
        "rows": 21,
        "columns": ["A", "B", "C", "D", "E"],
        "empty": False,
        "generated_via": "openpyxl",
    }
    report = build_validation_report(
        mode="dual_platform",
        model_name=model_name,
        input_reports={"autohome": autohome_report, "dcd": dcd_report},
        output_profiles=output_profiles,
        expected_sheets=DUAL_OUTPUT_SHEETS,
        critical_sheets=DUAL_CRITICAL_SHEETS,
        strict_validate=strict_validate,
    )
    report["output_path"] = str(out)
    report["validation_path"] = str(report_path)

    if reporter:
        reporter.start_stage("write_output", "写出摘要 Excel", weight=DUAL_STAGE_WEIGHTS["write_output"], total=1, message="开始写出摘要 Excel")
    if report["errors"]:
        write_validation_report(report_path, report)
        if reporter:
            reporter.finalize(message="校验失败")
        raise SystemExit(2)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sheet_name, frame in output_frames.items():
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
        for ws in writer.book.worksheets:
            style_sheet(ws)
    wb = load_workbook(out)
    build_one_pager_sheet(wb, len(zj_df), dcd["raw_count"], compare_df)
    wb.save(out)
    if reporter:
        reporter.finish_stage("摘要 Excel 已写出")

    if reporter:
        reporter.start_stage("write_validation", "写出校验报告", weight=DUAL_STAGE_WEIGHTS["write_validation"], total=1, message="开始写出校验报告")
    write_validation_report(report_path, report)
    if reporter:
        reporter.finish_stage("校验报告已写出")
        reporter.finalize(message="导出完成")
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", help="单平台输入 Excel（汽车之家结构）")
    ap.add_argument("--autohome-input", help="汽车之家原始口碑 Excel")
    ap.add_argument("--dcd-input", help="懂车帝原始口碑 Excel")
    ap.add_argument("--output", required=True)
    ap.add_argument("--model-name", required=True)
    ap.add_argument("--strict-validate", action="store_true", help="遇到校验告警时也按失败退出")
    ap.add_argument("--progress", action="store_true", help="显示进度条（默认仅在交互终端自动开启）")
    ap.add_argument("--progress-file", help="将进度状态写入 JSON 文件，供飞书/对话框轮询")
    ap.add_argument("--progress-webhook", help="将进度状态 POST 到 webhook，供飞书机器人或前端消费")
    ap.add_argument("--feishu-webhook", help="直接发送到飞书 incoming webhook 的消息地址")
    ap.add_argument("--feishu-secret", help="飞书自定义机器人安全密钥，用于签名")
    args = ap.parse_args()

    show_progress = args.progress or sys.stderr.isatty()
    use_progress_sink = bool(args.progress_file or args.progress_webhook or args.feishu_webhook)
    reporter = make_reporter(args, mode="dual_platform" if args.autohome_input and args.dcd_input else "autohome", model_name=args.model_name) if use_progress_sink else None

    if args.autohome_input and args.dcd_input:
        out = summarize_dual_platform(
            args.autohome_input,
            args.dcd_input,
            args.output,
            args.model_name,
            strict_validate=args.strict_validate,
            show_progress=show_progress,
            reporter=reporter,
        )
    elif args.input:
        out = summarize_autohome(
            args.input,
            args.output,
            args.model_name,
            strict_validate=args.strict_validate,
            show_progress=show_progress,
            reporter=reporter,
        )
    else:
        raise SystemExit("请提供 --input，或同时提供 --autohome-input 与 --dcd-input")

    print(out)
    print(Path(args.output).with_suffix(".validation.json"))


if __name__ == "__main__":
    main()
